#!/usr/bin/env python
# coding: utf-8

# ## 研究類各表（7張）

import urllib.request
import urllib.parse
import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import pandas as pd


url = ["https://udb.moe.edu.tw/download/udata/static_file/%E7%A0%941.%E5%AD%B8%E6%A0%A1%E6%89%BF%E6%8E%A5%E5%90%84%E5%96%AE%E4%BD%8D%E8%B3%87%E5%8A%A9%E3%80%8C%E5%90%84%E9%A1%9E%E8%A8%88%E7%95%AB%E7%B6%93%E8%B2%BB%E3%80%8D%E5%8F%8A%E5%85%B6%E6%AF%8F%E5%B8%AB%E5%B9%B3%E5%9D%87%E6%89%BF%E6%8E%A5%E9%87%91%E9%A1%8D-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E7%A0%942.%E5%AD%B8%E6%A0%A1%E6%89%BF%E6%8E%A5%E5%90%84%E5%96%AE%E4%BD%8D%E8%B3%87%E5%8A%A9%E3%80%8C%E7%94%A2%E5%AD%B8%E5%90%88%E4%BD%9C%E3%80%8D%E8%A8%88%E7%95%AB%E7%B6%93%E8%B2%BB%E5%8F%8A%E5%85%B6%E6%AF%8F%E5%B8%AB%E5%B9%B3%E5%9D%87%E6%89%BF%E6%8E%A5%E9%87%91%E9%A1%8D-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E7%A0%943.%E5%AD%B8%E6%A0%A1%E7%94%B3%E8%AB%8B%E5%B0%88%E5%88%A9%E3%80%81%E6%96%B0%E5%93%81%E7%A8%AE%E5%8F%8A%E6%8E%88%E6%AC%8A%E4%BB%B6%E6%95%B8-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E7%A0%944.%E5%AD%B8%E6%A0%A1%E5%90%84%E7%A8%AE%E6%99%BA%E6%85%A7%E8%B2%A1%E7%94%A2%E6%AC%8A%E8%A1%8D%E7%94%9F%E9%81%8B%E7%94%A8%E7%B8%BD%E9%87%91%E9%A1%8D-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv",
       "https://udb.moe.edu.tw/download/udata/static_file/%E7%A0%945.%E5%B0%88%E4%BB%BB%E6%95%99%E5%B8%AB%E7%8D%B2%E8%B3%87%E5%8A%A9%E3%80%8C%E5%AD%B8%E8%A1%93%E7%A0%94%E7%A9%B6%E3%80%8D%E8%A8%88%E7%95%AB%E7%B6%93%E8%B2%BB%E5%8F%8A%E5%85%B6%E6%AF%8F%E5%B8%AB%E5%B9%B3%E5%9D%87%E6%89%BF%E6%8E%A5%E9%87%91%E9%A1%8D-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E7%A0%945.%E5%B0%88%E4%BB%BB%E6%95%99%E5%B8%AB%E7%8D%B2%E8%B3%87%E5%8A%A9%E3%80%8C%E5%AD%B8%E8%A1%93%E7%A0%94%E7%A9%B6%E3%80%8D%E8%A8%88%E7%95%AB%E7%B6%93%E8%B2%BB%E5%8F%8A%E5%85%B6%E6%AF%8F%E5%B8%AB%E5%B9%B3%E5%9D%87%E6%89%BF%E6%8E%A5%E9%87%91%E9%A1%8D-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E7%A0%946.%E5%AD%B8%E6%A0%A1%E8%A1%8D%E7%94%9F%E4%BC%81%E6%A5%AD-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E7%A0%947.%E5%AD%B8%E6%A0%A1%E5%90%88%E4%BD%9C%E4%BC%81%E6%A5%AD%E6%96%B0%E4%BA%8B%E6%A5%AD%E9%83%A8%E9%96%80-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv"]
local_directory = '/Users/issac/Desktop/Research/'
filenames = []

m = len(url)
for i in range(m):
    filename_i = os.path.basename(url[i])
    filename_i = urllib.parse.unquote(filename_i, encoding='utf-8')
    local_file_path = os.path.join(local_directory, filename_i)
    urllib.request.urlretrieve(url[i], local_file_path)
    filenames.append(filename_i)


# Read the original csv sheets

existing = f'{local_directory}Research.xlsx'
if os.path.isfile(existing):
    os.remove(existing)
Research = openpyxl.Workbook()
Research.save(existing)
for fn in filenames:
    csv_fn = pd.read_csv(f'{local_directory}{fn}')
    # Get all the original names of columns
    sheet_fn = pd.DataFrame({'原始欄位名稱': list(csv_fn.columns)})
    columns_to_add = ['欄位名稱(中)', '欄位名稱(英)', '中文值標籤', '序號', '型態', '簡易說明', '備註', '資料變動說明',
                      '處理後刪除', '新增欄位', '資料表名稱', '檢誤邏輯', '資料庫簡介', '資料涵蓋學年度', '欄位勾選', '年份選取', '層級']
    sheet_fn = sheet_fn.reindex(
        columns=[*sheet_fn.columns, *columns_to_add], fill_value=np.nan)
    # Fill in 層級
    unit = '以「校'
    if unit in fn:
        sheet_fn['層級'] = '校'
    else:
        sheet_fn['層級'] = '系'
    # Fill in 資料涵蓋學年度 (some columns might be NaN for some years)
    sn = list(csv_fn.columns)
    for s in sn:  # 可能要事前確認一致變項名稱，事前再建立一個data過的csv來iterate
        csv_s = csv_fn[csv_fn[s] != '...']
        if '學年度' in csv_s:
            years = csv_s['學年度'].unique().astype(str).tolist()
            sheet_fn.loc[sheet_fn['原始欄位名稱'] == s, '資料涵蓋學年度'] = '; '.join(years)
        else:
            years = csv_s['年度'].unique().astype(str).tolist()
            sheet_fn.loc[sheet_fn['原始欄位名稱'] == s, '資料涵蓋學年度'] = '; '.join(years)
    # Fill in 型態
        sheet_fn.loc[sheet_fn['原始欄位名稱'] == s, '型態'] = csv_fn[s].dtype
    # Fill in 中文值標籤
    for s in sn:
        vallabs = csv_fn[s].unique().astype(str).tolist()
        sheet_fn.loc[(sheet_fn['原始欄位名稱'] == s) & (len(vallabs) < 10) & (
            csv_fn[s].dtype == object), '中文值標籤'] = '; '.join(vallabs)
    # Fill in 資料表名稱
    sheet_fn['資料表名稱'] = fn
    # Reorder
    sheet_fn = sheet_fn.loc[:, ['欄位名稱(中)', '原始欄位名稱', '欄位名稱(英)', '中文值標籤', '序號', '型態', '簡易說明', '備註',
                                '資料變動說明', '處理後刪除', '新增欄位', '資料表名稱', '檢誤邏輯', '資料庫簡介', '資料涵蓋學年度', '欄位勾選', '年份選取', '層級']]
    # Export
    book = load_workbook(f'{local_directory}Research.xlsx')
    writer = pd.ExcelWriter(
        f'{local_directory}Research.xlsx', engine='openpyxl')
    writer.book = book
    sheet_fn.to_excel(writer, sheet_name=fn[:5])
    writer.save()
    writer.close()
# Delete the blank sheet
Research = openpyxl.load_workbook(f'{local_directory}Research.xlsx')
Research.remove(Research['Sheet'])
Research.save(f'{local_directory}Research.xlsx')


# ## 學生類各表（36張）


url = ["https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B81-1.%E6%AD%A3%E5%BC%8F%E5%AD%B8%E7%B1%8D%E5%9C%A8%E5%AD%B8%E5%AD%B8%E7%94%9F%E4%BA%BA%E6%95%B8-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B81-2.%E6%AD%A3%E5%BC%8F%E5%AD%B8%E7%B1%8D%E5%9C%A8%E5%AD%B8%E5%AD%B8%E7%94%9F%E4%BA%BA%E6%95%B8-%E4%BB%A5%E3%80%8C%E6%A0%A1(%E5%90%AB%E5%AD%B8%E5%88%B6%E7%8F%AD%E5%88%A5)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B82-1.%E7%95%A2%E6%A5%AD%E7%94%9F%E6%95%B8%E5%8F%8A%E5%85%B6%E5%8F%96%E5%BE%97%E8%BC%94%E7%B3%BB%E3%80%81%E9%9B%99%E4%B8%BB%E4%BF%AE%E8%B3%87%E6%A0%BC%E4%BA%BA%E6%95%B8-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B82-2.%E7%95%A2%E6%A5%AD%E7%94%9F%E6%95%B8%E5%8F%8A%E5%85%B6%E5%8F%96%E5%BE%97%E8%BC%94%E7%B3%BB%E3%80%81%E9%9B%99%E4%B8%BB%E4%BF%AE%E8%B3%87%E6%A0%BC%E4%BA%BA%E6%95%B8-%E4%BB%A5%E3%80%8C%E6%A0%A1(%E5%90%AB%E5%AD%B8%E5%88%B6%E7%8F%AD%E5%88%A5)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B82-3.%E7%95%A2%E6%A5%AD%E6%8E%88%E4%BA%88%E5%AD%B8%E4%BD%8D%E4%B9%8B%E4%B8%AD%E8%8B%B1%E6%96%87%E5%90%8D%E7%A8%B1%E5%8F%8A%E4%BA%BA%E6%95%B8%E8%A1%A8-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B82-4.%E7%95%A2%E6%A5%AD%E7%A2%A9%E3%80%81%E5%8D%9A%E5%A3%AB%E5%AD%B8%E4%BD%8D%E8%AB%96%E6%96%87%E8%B3%87%E6%96%99-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B82-5.%E7%95%A2%E6%A5%AD%E7%A2%A9%E3%80%81%E5%8D%9A%E5%A3%AB%E5%AD%B8%E4%BD%8D%E8%AB%96%E6%96%87%E8%B3%87%E6%96%99-%E4%BB%A5%E3%80%8C%E6%A0%A1(%E5%90%AB%E5%AD%B8%E5%88%B6%E7%8F%AD%E5%88%A5)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B83-1.%E5%A2%83%E5%A4%96%E5%AD%B8%E4%BD%8D%E7%94%9F%E6%95%B8%E5%8F%8A%E5%85%B6%E5%9C%A8%E5%AD%B8%E6%AF%94%E7%8E%87-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B83-2.%E5%A4%96%E5%9C%8B%E5%AD%B8%E7%94%9F%E6%95%B8%E5%8F%8A%E5%85%B6%E5%9C%A8%E5%AD%B8%E6%AF%94%E7%8E%87-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B83-3.%E5%83%91%E7%94%9F%E3%80%81%E6%B8%AF%E6%BE%B3%E7%94%9F%E6%95%B8%E5%8F%8A%E5%85%B6%E5%9C%A8%E5%AD%B8%E6%AF%94%E7%8E%87-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B83-4.%E5%A4%A7%E9%99%B8%E5%9C%B0%E5%8D%80%E4%BE%86%E8%87%BA%E5%AD%B8%E4%BD%8D%E7%94%9F%E6%95%B8%E5%8F%8A%E5%85%B6%E5%9C%A8%E5%AD%B8%E6%AF%94%E7%8E%87-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B83-5.%E7%95%A2%E6%A5%AD%E5%A2%83%E5%A4%96%E7%94%9F%E6%95%B8-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B83-6.%E7%95%A2%E6%A5%AD%E5%A2%83%E5%A4%96%E7%94%9F%E6%95%B8-%E4%BB%A5%E3%80%8C%E6%A0%A1(%E5%90%AB%E5%AD%B8%E5%88%B6%E7%8F%AD%E5%88%A5)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B83-7.%E5%A2%83%E5%A4%96%E5%AD%B8%E4%BD%8D%E7%94%9F%E6%95%B8(%E4%BE%9D%E8%BA%AB%E5%88%86%E9%A1%9E%E5%88%A5%E5%91%88%E7%8F%BE)-%20%E4%BB%A5%E3%80%8C%E6%A0%A1(%E5%90%AB%E5%AD%B8%E5%88%B6%E7%8F%AD%E5%88%A5)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B84.%E6%97%A5%E9%96%93%E5%AD%B8%E5%A3%AB%E7%8F%AD%E4%BB%A5%E4%B8%8B%E7%94%B3%E8%AB%8B%E5%BB%B6%E9%95%B7%E4%BF%AE%E6%A5%AD%E5%B9%B4%E9%99%90%E4%B9%8B%E5%BB%B6%E4%BF%AE%E7%94%9F%E6%95%B8-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B85.%E6%97%A5%E9%96%93%E5%AD%B8%E5%88%B6%E6%9C%AC%E5%9C%8B%E5%AD%B8%E7%94%9F%E5%87%BA%E5%9C%8B%E9%80%B2%E4%BF%AE%E4%BA%A4%E6%B5%81%E4%BA%BA%E6%95%B8-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B86.%E4%BF%AE%E8%AE%80%E6%A0%A1%E9%9A%9B%E9%81%B8%E8%AA%B2%E4%BA%BA%E6%AC%A1-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv",
       "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B87.%E4%BF%AE%E8%AE%80%E8%BC%94%E7%B3%BB%E4%BA%BA%E6%AC%A1-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B88.%E4%BF%AE%E8%AE%80%E9%9B%99%E4%B8%BB%E4%BF%AE%E4%BA%BA%E6%AC%A1-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B89.%E5%AD%B8%E6%A0%A1%E9%9B%99%E8%81%AF%E5%90%88%E4%BD%9C%E6%A1%88%E4%BB%B6%E6%95%B8-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B810.%E6%97%A5%E9%96%93%E5%AD%B8%E5%88%B6%E4%BF%AE%E8%AE%80%E9%9B%99%E8%81%AF%E5%AD%B8%E5%88%B6%E5%AD%B8%E7%94%9F%E6%95%B8-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B811.%E5%AD%B8%E6%A0%A1%E8%BE%A6%E7%90%86%E5%9C%8B%E9%9A%9B%E5%90%88%E4%BD%9C%E8%88%87%E4%BA%A4%E6%B5%81%E8%B3%87%E6%96%99%E8%A1%A8-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B812-1.%E6%96%B0%E7%94%9F(%E5%90%AB%E5%A2%83%E5%A4%96%E7%94%9F)%E8%A8%BB%E5%86%8A%E7%8E%87-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B812-2.%E6%96%B0%E7%94%9F(%E5%90%AB%E5%A2%83%E5%A4%96%E7%94%9F)%E8%A8%BB%E5%86%8A%E7%8E%87-%E4%BB%A5%E3%80%8C%E6%A0%A1(%E5%90%AB%E5%AD%B8%E5%88%B6%E7%8F%AD%E5%88%A5)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B812-3.%E6%96%B0%E7%94%9F(%E5%90%AB%E5%A2%83%E5%A4%96%E7%94%9F)%E8%A8%BB%E5%86%8A%E7%8E%87-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B813-1.%E6%96%BC%E5%AD%B8%E5%B9%B4%E5%BA%95%E8%99%95%E6%96%BC%E4%BC%91%E5%AD%B8%E7%8B%80%E6%85%8B%E4%B9%8B%E4%BA%BA%E6%95%B8-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B813-2.%E6%96%BC%E5%AD%B8%E5%B9%B4%E5%BA%95%E8%99%95%E6%96%BC%E4%BC%91%E5%AD%B8%E7%8B%80%E6%85%8B%E4%B9%8B%E4%BA%BA%E6%95%B8-%E4%BB%A5%E3%80%8C%E6%A0%A1(%E5%90%AB%E5%AD%B8%E5%88%B6%E7%8F%AD%E5%88%A5)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B814-1.%E9%80%80%E5%AD%B8%E4%BA%BA%E6%95%B8-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B814-2.%E9%80%80%E5%AD%B8%E4%BA%BA%E6%95%B8-%E4%BB%A5%E3%80%8C%E6%A0%A1(%E5%90%AB%E5%AD%B8%E5%88%B6%E7%8F%AD%E5%88%A5)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B815.%E5%9C%A8%E5%AD%B8%E5%AD%B8%E7%94%9F%E5%8F%83%E8%88%87%E7%AB%B6%E8%B3%BD%E3%80%81%E8%AB%96%E6%96%87%E5%87%BA%E7%89%88%E7%AD%89%E6%88%90%E6%95%88-%E4%BB%A5%E3%80%8C%E6%A0%A1(%E5%90%AB%E5%AD%B8%E5%88%B6%E7%8F%AD%E5%88%A5)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B816.%E5%AD%B8%E5%A3%AB%E7%8F%AD%E4%BB%A5%E4%B8%8B%E5%B0%B1%E5%AD%B8%E7%A9%A9%E5%AE%9A%E7%8E%87-%E4%BB%A5%E3%80%8C%E6%A0%A1(%E5%90%AB%E5%AD%B8%E5%88%B6%E7%8F%AD%E5%88%A5)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B817.%E7%8D%8E%E5%8A%A9%E7%94%9F%E5%8F%8A%E5%8B%9E%E5%83%B1%E5%9E%8B%E5%AD%B8%E7%94%9F%E5%85%BC%E4%BB%BB%E5%8A%A9%E7%90%86%E4%BA%BA%E6%95%B8-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B818.%E5%AD%B8%E6%A0%A1%E5%9F%B7%E8%A1%8C%E5%B7%A5%E8%AE%80%E5%8F%8A%E7%94%9F%E6%B4%BB%E5%8A%A9%E5%AD%B8%E9%87%91%E6%83%85%E5%BD%A2-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B819.%E5%85%BC%E4%BB%BB%E5%8A%A9%E7%90%86%E5%B9%B3%E5%9D%87%E6%AF%8F%E6%9C%88%E6%94%AF%E7%B5%A6%E9%87%91%E9%A1%8D%E4%BA%BA%E6%95%B8%E7%B5%B1%E8%A8%88%E8%A1%A8-%E4%BB%A5%E3%80%8C%E6%A0%A1%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B820.%E5%AD%B8%E7%94%9F%E5%AF%A6%E7%BF%92%E6%83%85%E5%BD%A2-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv", "https://udb.moe.edu.tw/download/udata/static_file/%E5%AD%B821.%E5%AD%B8%E7%94%9F%E5%AF%A6%E7%BF%92%E6%A9%9F%E6%A7%8B%E5%8F%8A%E6%AC%8A%E7%9B%8A%E4%BF%9D%E9%9A%9C-%E4%BB%A5%E3%80%8C%E7%B3%BB(%E6%89%80)%E3%80%8D%E7%B5%B1%E8%A8%88.csv"]
local_directory = '/Users/issac/Desktop/Student/'
filenames = []

m = len(url)
for i in range(m):
    filename_i = os.path.basename(url[i])
    filename_i = urllib.parse.unquote(filename_i, encoding='utf-8')
    local_file_path = os.path.join(local_directory, filename_i)
    urllib.request.urlretrieve(url[i], local_file_path)
    filenames.append(filename_i)


# Read the original csv sheets

existing = f'{local_directory}Student.xlsx'
if os.path.isfile(existing):
    os.remove(existing)
Student = openpyxl.Workbook()
Student.save(existing)
for fn in filenames:
    csv_fn = pd.read_csv(f'{local_directory}{fn}')
    # Get all the original names of columns
    sheet_fn = pd.DataFrame({'原始欄位名稱': list(csv_fn.columns)})
    columns_to_add = ['欄位名稱(中)', '欄位名稱(英)', '中文值標籤', '序號', '型態', '簡易說明', '備註', '資料變動說明',
                      '處理後刪除', '新增欄位', '資料表名稱', '檢誤邏輯', '資料庫簡介', '資料涵蓋學年度', '欄位勾選', '年份選取', '層級']
    sheet_fn = sheet_fn.reindex(
        columns=[*sheet_fn.columns, *columns_to_add], fill_value=np.nan)
    # Fill in 層級
    unit = '以「校'
    if unit in fn:
        sheet_fn['層級'] = '校'
    else:
        sheet_fn['層級'] = '系'
    # Fill in 資料涵蓋學年度 (some columns might be NaN for some years)
    sn = list(csv_fn.columns)
    for s in sn:
        csv_s = csv_fn[csv_fn[s] != '...']
        years = csv_s['學年度'].unique().astype(str).tolist()
        sheet_fn.loc[sheet_fn['原始欄位名稱'] == s, '資料涵蓋學年度'] = '; '.join(years)
    # Fill in 型態
        sheet_fn.loc[sheet_fn['原始欄位名稱'] == s, '型態'] = csv_fn[s].dtype
    # Fill in 中文值標籤
    for s in sn:
        vallabs = csv_fn[s].unique().astype(str).tolist()
        sheet_fn.loc[(sheet_fn['原始欄位名稱'] == s) & (len(vallabs) < 10) & (
            csv_fn[s].dtype == object), '中文值標籤'] = '; '.join(vallabs)
    # Fill in 資料表名稱
    sheet_fn['資料表名稱'] = fn
    # Reorder
    sheet_fn = sheet_fn.loc[:, ['欄位名稱(中)', '原始欄位名稱', '欄位名稱(英)', '中文值標籤', '序號', '型態', '簡易說明', '備註',
                                '資料變動說明', '處理後刪除', '新增欄位', '資料表名稱', '檢誤邏輯', '資料庫簡介', '資料涵蓋學年度', '欄位勾選', '年份選取', '層級']]
    # Export
    book = load_workbook(f'{local_directory}Student.xlsx')
    writer = pd.ExcelWriter(
        f'{local_directory}Student.xlsx', engine='openpyxl')
    writer.book = book
    sheet_fn.to_excel(writer, sheet_name=fn[:5])
    writer.save()
    writer.close()
# Delete the blank sheet
Student = openpyxl.load_workbook(f'{local_directory}Student.xlsx')
Student.remove(Student['Sheet'])
Student.save(f'{local_directory}Student.xlsx')
